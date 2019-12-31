import requests
import json
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

'''
全局变量
'''
print(''' 
=@@@@@@@@@@@@@@@@@@                            =@@@@@@@@@@@@@@@\*                              =@@@@^                                            
=@@@@@@@@@@@@@@@@@/                            =@@@@@@@@@@@@@@@@@\                             =@@@@^                                            
=@@@@                                          =@@@@^         \@@@@                            =@@@@^                                            
=@@@@                                          =@@@@^          @@@@^                           =@@@@^                                            
=@@@@                                          =@@@@^          @@@@@                           =@@@@^                                            
=@@@@                                          =@@@@^          @@@@^                           =@@@@^                                            
=@@@@]]]]]]]]]]]]]                             =@@@@^         =@@@@`                           =@@@@^                                            
=@@@@@@@@@@@@@@@@@                             =@@@@\]]]]]]]@@@@@@`                            =@@@@^                                            
=@@@@[[[[[[[[[[[[`                             =@@@@@@@@@@@@@@@@`                              =@@@@^                                            
=@@@@                                          =@@@@/[[[[[[[[                                  =@@@@^                                            
=@@@@                                          =@@@@^                                          =@@@@^                                            
=@@@@                                          =@@@@^                                          =@@@@^                                            
=@@@@                                          =@@@@^                                          =@@@@^                                            
=@@@@                                          =@@@@^                                          =@@@@^                                            
=@@@@@@@@@@@@@@@@@@@                           =@@@@^                                          =@@@@@@@@@@@@@@@@@@@^                             
=@@@@@@@@@@@@@@@@@@@                           =@@@@^                                          =@@@@@@@@@@@@@@@@@@@^                                                                                                                                                                                                                                                                                         
                                              
                   \@@^  *@@@^   @@@`    /@@@@@@@@`      @@@ ]@@@@@@@`       ]@@@@@@@/@@@@   @@@     ,@@@/       /@@@@@@@@`           @@@^       
                   =@@\  =@@@@  =@@/   =@@@[   *\@@@     @@@@@[   ,@@@^     @@@`   \@@^      @@@   ,@@@/       =@@@[   *\@@@          @@@^       
                    @@@  @@^@@^ =@@^   [[[`      @@@     @@@/      =@@@    =@@^     @@@      @@@ ,@@@/         [[[`      @@@          @@@^       
                    =@@^,@@ \@\ @@@      ,]@@@@@@@@@     @@@       =@@@    ,@@@    =@@/      @@@@@@@@@`          ,]@@@@@@@@@          @@@^       
                    @@@=@/ =@@=@@^    /@@@/[[*  @@@     @@@       =@@@     ,@@@@@@@@/       @@@@`  @@@`       /@@@/[[*  @@@          @@@^       
                    =@@@@^  @@@@@    =@@@      ,@@@     @@@       =@@@    @@@  [[`          @@@     \@@\     =@@@      ,@@@          @@@^       
                    *@@@@   =@@@^    =@@@\* ,]@@@@@*    @@@       =@@@    \@@@@@@@@@]]      @@@      =@@\    =@@@\* ,]@@@@@*         @@@^       
                    \@@^   ,@@@*     ,\@@@@@@/`\@@\    @@@       =@@@    ,@@/[@@@@@@@@\    @@@       =@@@`   ,\@@@@@@/`\@@\         @@@^       
                                                                        /@@         =@@^                                                       
                                                                        ,@@@@@]]]]@@@@`                                                        
                                                                            *[[[[[[[*                                                           
                                                                                                                                                                                                                                
''')
print('----------------------------------------------------------------------------')
print('------------------------------欢迎来到expl收文批处理小程序------------------')
print('----------------------------------------------------------------------------')
print('''
演示案例
请输入OA账号：w**kai
请输入OA密码：123456
请输入需要处理收文的开始时间：2019-11-01
请输入需要处理收文的结束时间：2019-12-31
''')
list1 = ['请输入OA账号：','请输入OA密码：','请输入需要处理收文的开始时间：','请输入需要处理收文的结束时间：']
print(list1[0])
username=input()
print(list1[1])
password=input()
print(list1[2])
start_time=input()
print(list1[3])
end_time=input()






# start_time = '2019-11-01'
# end_time = '2019-12-31'
# global jishu
jishu = 0


'''
1.获取标识JSESSIONID
'''
url = "http://oa.expl.cn/seeyon/"
response = requests.get(url)
# print(type(response.cookies))
cookies = requests.utils.dict_from_cookiejar(response.cookies)
# print(cookies['JSESSIONID'])

# print(type(cookies))


# 
'''
2.输入账号密码和拼接JSESSIONID
'''

login_headers = {
'Host': 'oa.expl.cn',
'Cache-Control': 'max-age=0',
'Origin': 'http://oa.expl.cn',
'Upgrade-Insecure-Requests': '1',
'Content-Type': 'application/x-www-form-urlencoded',
'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36',
'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
'Referer': 'http://oa.expl.cn/seeyon/index.jsp',
'Accept-Language': 'zh-CN,zh;q=0.9',
'Cookie': 'login.locale=zh_CN'
}
login_headers['Cookie']='JSESSIONID='+cookies['JSESSIONID']+';login.locale=zh_CN'
# print(login_headers)

login_data={}
# UserAgentFrom=pc&login.username=kai&login.password=0123
login_data['UserAgentFrom'] = 'pc'
login_data['login.username'] = username
login_data['login.password'] = password
# login_date[''] = 
# login_date[''] = 
r = requests.post('http://oa.expl.cn/seeyon/login/proxy', data=login_data, headers=login_headers)
# ret_dict = json.loads(r.text)
# print(login_data)
# print(r.text)
'''
3.验证过后，查询收文-目录
3.1 查看目录的记录count
---
'''
params = {}
mulu_headers = {
'Host': 'oa.expl.cn',
'Cache-Control': 'max-age=0',
'Origin': 'http://oa.expl.cn',
'Upgrade-Insecure-Requests': '1',
'Content-Type': 'application/x-www-form-urlencoded',
'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36',
'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
'Referer': 'http://oa.expl.cn/seeyon/main.do?method=left&fromPortal=false',
'Accept-Language': 'zh-CN,zh;q=0.9'
}

mulu_headers['Cookie']='JSESSIONID='+cookies['JSESSIONID']+';login.locale=zh_CN'

# r2 = requests.get('http://oa.expl.cn/seeyon/main.do?method=morePending4App&app=Edoc', params=params,headers=login_headers)
r2 = requests.get('http://oa.expl.cn/seeyon/edocController.do?method=listDone&appName=4&edocType=1&edocMarkValue=&edocInMarkValue=&condition=createDate&textfield=%s&textfield1=%s'%(start_time,end_time), params=params,headers=login_headers)
# /seeyon/edocController.do?method=listDone&edocType=1&condition=&textfield=&textfield1=
# print(r2.text)
'''
查询目录-得到当页公文所有value值组成一个list 特征码（px;"><input type="checkbox" name='id' value="4779076869180116767" false category="）
'''

# pattern_value = '<input type="checkbox" name=\'id\' value="([\s\S]*?)"'
# pattern_value = '<input type=\'checkbox\' name=\'id\' value="([\s\S]*?)" '
pattern_count = '条/共([\s\S]*?)条记录 '
# 得到发起时间这个条件筛选count条数
value_count = re.findall(pattern_count,r2.text)
# print(value_count[0])

# value_list = re.findall(pattern_value,r2.text)
'''
3.验证过后，查询收文-目录
3.2 查看目录的id list
---
'''
params = {}
mulu_headers = {
'Host': 'oa.expl.cn',
'Cache-Control': 'max-age=0',
'Origin': 'http://oa.expl.cn',
'Upgrade-Insecure-Requests': '1',
'Content-Type': 'application/x-www-form-urlencoded',
'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36',
'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
'Referer': 'http://oa.expl.cn/seeyon/main.do?method=left&fromPortal=false',
'Accept-Language': 'zh-CN,zh;q=0.9'
}

mulu_headers['Cookie']='JSESSIONID='+cookies['JSESSIONID']+';login.locale=zh_CN'

# r2 = requests.get('http://oa.expl.cn/seeyon/main.do?method=morePending4App&app=Edoc', params=params,headers=login_headers)
r2 = requests.get('http://oa.expl.cn/seeyon/edocController.do?method=listDone&appName=4&edocType=1&edocMarkValue=&edocInMarkValue=&condition=createDate&textfield=%s&textfield1=%s&page=1&pageSize=%s'%(start_time,end_time,value_count[0]), params=params,headers=login_headers)
# /seeyon/edocController.do?method=listDone&edocType=1&condition=&textfield=&textfield1=
# print(r2.text)
'''
查询目录-得到当页公文所有value值组成一个list 特征码（px;"><input type="checkbox" name='id' value="4779076869180116767" false category="）
'''

# pattern_value = '<input type="checkbox" name=\'id\' value="([\s\S]*?)"'
pattern_value = '<input type=\'checkbox\' name=\'id\' value="([\s\S]*?)" '
value_list = re.findall(pattern_value,r2.text)
# print(value_list)
# print(type(value_list))

# 
'''
4.查看公文内容
'''
def read_gongwen(gognwenid):
    params = {}
    gongwen_headers = {
    'Host': 'oa.expl.cn',
    'Cache-Control': 'max-age=0',
    'Origin': 'http://oa.expl.cn',
    'Upgrade-Insecure-Requests': '1',
    'Content-Type': 'application/x-www-form-urlencoded',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'Referer': 'http://oa.expl.cn/seeyon/main.do?method=left&fromPortal=false',
    'Accept-Language': 'zh-CN,zh;q=0.9'
    }

    gongwen_headers['Cookie']='JSESSIONID='+cookies['JSESSIONID']+';login.locale=zh_CN'
    # gognwenid = '-4720121956674482570'
    r3 = requests.get('http://oa.expl.cn/seeyon/edocController.do?method=getContent&summaryId=%s&affairId=-2103818690292857492&from=Done&openFrom=&lenPotent=&docId=&docResId=&canUploadRel=&canUploadAttachment=&position=&firstPDFId='%gognwenid, params=params,headers=gongwen_headers)
    # print(r3.text)
    # 
    '''
    匹配规则
    '''
    # root_pattern = '<textarea id="xml" cols="40" rows="10">([\s\S]*?)</textarea>'
    pattern_title = '<my:subject>([\s\S]*?)</my:subject>'
    pattern_num = '<my:doc_mark>([\s\S]*?)</my:doc_mark>'
    pattern_time = '<my:createdate>([\s\S]*?)</my:createdate>'
    # pattern_type2 = '<Input display="([\s\S]*?)" value=".*" select="true" />'
    # pattern_type = '<FieldInput name="my:secret_level" type="select" access="edit" allowprint="true" allowtransmit="true">([\s\S]*?)/>'
    pattern_type = '<my:secret_level>([\s\S]*?)</my:secret_level>'

    pattern_neibu = '>([\s\S]*?)<'

    pattern_opinion1 = '\["opinion1","<([\s\S]*?)"]'
    pattern_opinion2 = '\["niban","<([\s\S]*?)"]'

    wenjian_title = re.findall(pattern_title,r3.text)
    wenjian_num = re.findall(pattern_num,r3.text)
    wenjian_time = re.findall(pattern_time,r3.text)
    wenjian_type = re.findall(pattern_type,r3.text)
    # wenjian_neibu = 
    # wenjian_end = re.findall(pattern_end,r3.text)
    # print(wenjian_type[0])
    '''
    处理公文类型（特例）
    '''
    if wenjian_type[0] == '1':
        wenjian_type_value = '普通'
    elif wenjian_type[0] == '5':
        wenjian_type_value = '内部资料'
    elif wenjian_type[0] == '6':
        wenjian_type_value = '核心商密'
    elif wenjian_type[0] == '7':
        wenjian_type_value = '普通商密'
    elif wenjian_type[0] == '8':
        wenjian_type_value = '内部事项'
    else:
        wenjian_type_value = '未能识别：'+wenjian_type[0]+'。请联系开发添加!'
    

    # for i in wenjian_type:
    #     wenjian_type = i
    # print(wenjian_type)
    # 再筛选一次
    # wenjian_type = re.findall(pattern_type2,wenjian_type)
    # print(wenjian_type)
    # wenjian_title = re.findall(root_html,pattern_title)

    '''
    处理意见（特例）
    '''
    wenjian_opinion1 = re.findall(pattern_opinion1,r3.text)
    wenjian_opinion2 = re.findall(pattern_opinion2,r3.text)
    
    '''
    <Input display="普通" value="1" select="true" />
    <Input display="内部资料" value="5"/>
    <Input display="核心商密" value="6"/
    ><Input display="普通商密" value="7"/>
    <Input display="内部事项" value="8"/>
    '''
    '''
    处理条数
    ---时间筛选   /seeyon/edocController.do?method=listDone&appName=4&edocType=1&edocMarkValue=&edocInMarkValue=&condition=createDate&textfield=2019-11-01&textfield1=2019-12-31
    ------筛选匹配 count 
    ---条数控制   &page=1&pageSize=3

    '''
    # if wenjian_type=='':
    #     wenjian_type = ''
    # ([\s\S]*?)

    print(wenjian_title[0])
    print(wenjian_num[0])
    print(wenjian_time[0])
    print(wenjian_type_value)
    # print(wenjian_opinion1[0])
    # print(re.findall(pattern_neibu,wenjian_opinion1[0])[0])
    # print(re.findall(pattern_neibu,wenjian_opinion1[0])[1])
    nibian =   str(re.findall(pattern_neibu,wenjian_opinion1[0])[0])
    nibian_p = str(re.findall(pattern_neibu,wenjian_opinion1[0])[1])

    # print(wenjian_opinion2[0])
    option =   str(re.findall(pattern_neibu,wenjian_opinion2[0])[0])
    option_p = str(re.findall(pattern_neibu,wenjian_opinion2[0])[1])
    # print(re.findall(pattern_neibu,wenjian_opinion2[0])[0])
    # print(re.findall(pattern_neibu,wenjian_opinion2[0])[1])
    '''
    写xls
    '''
    # ws.cell('123')
    values = [wenjian_num[0],'易普力',wenjian_title[0],wenjian_time[0],wenjian_type_value,nibian,nibian_p,option,option_p]

    nrows = ws.max_row # 获得行数
    # ncolumns = ws.max_column # 获得列数
    nrows = nrows + 1
    ws.cell(nrows,1).value = values[0]
    ws.cell(nrows,2).value = values[1]
    ws.cell(nrows,3).value = values[2]
    ws.cell(nrows,4).value = values[3]
    ws.cell(nrows,5).value = values[4]
    ws.cell(nrows,6).value = values[5]
    ws.cell(nrows,7).value = values[6]
    ws.cell(nrows,8).value = values[7]
    ws.cell(nrows,9).value = values[8]



# main
'''
首先打开创建一个execl
'''
# 设置文件 mingc
addr = "C:\\%s.xlsx"%(datetime.datetime.now().strftime("%H%M%S"))
wb = Workbook()
# 激活 worksheet
ws = wb.active

ws.title = 'test_sheet1'

# nrows = ws.max_row # 获得行数
# ncolumns = ws.max_column # 获得列数

ws['A1'] = '公文文号'
ws['B1'] = '来文单位'
ws['C1'] = '文件标题'
ws['D1'] = '收文时间'
ws['E1'] = '文件密集'
ws['F1'] = '拟办意见'
ws['G1'] = '拟办处理人'
ws['H1'] = '批示意见'
ws['I1'] = '批示处理人'

for  i  in  value_list:
    # print(i)

    jishu = jishu+1
    print('开始-----------%s----------------------'%jishu)
    read_gongwen(i)
    print('-----------------------------------结束')
print('共处理:'+value_count[0])
wb.save(addr)